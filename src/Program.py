import requests
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl import Workbook
from datetime import date
import os
import time
from time import gmtime
from time import strftime
import threading
import Amascotados
import Biscoitinho
import Petcity
import Meganimal
import Miscota
import Newpetclub
import Petcity
import Petness
import Petsonic
import PiensosYMascotas
import PlanetaHuerto
import TiendaAnimal
import Wecanimal
import Kiwoko
import Biscoitinho
import Goldpet
import Zooshop

threads = []

ffthreads = []

runnablePetPricesList = [
                    'Meganimal.run',
                    'Miscota.run',
                    'Newpetclub.run',
                    'Petcity.run',
                    'Petness.run',
                    'Petsonic.run',
                    'PiensosYMascotas.run',
                    'PlanetaHuerto.run',
                    'TiendaAnimal.run',
                    'Wecanimal.run',
                    ]

notRunnablePetPricesList = [
                        'Amascotados',
                        'Kiwoko',
                        'Biscoitinho',
                        'Goldpet',
                        'Zooshop',
                        ]



filesList = ['amascotadosES',
             'amascotadosPT',
             'biscoitinho',
             'goldpet',
             'KiwokoEs',
             'KiwokoPT',
             'meganimal',
             'miscotaES',
             'miscotaPT',
             'newPetClub',
             'petcity',
             'petness',
             'petsonic',
             'piensosYMascotas',
             'planetahuertoEs',
             'planetahuertoPt',
             'tiendaAnimalES',
             'tiendaAnimalPT',
             'wecanimal',
             'zooshop',
             ]

path2 = 'teste.xlsx'

testfinalPath = 'C:\\<path>>\\<file>.xlsx'
defaultPath = 'C:\\<path>>\\'
fileExtension = '.xlsx'

today = date.today()
d = today.strftime("%Y%B%d")

def xlsxJoiner():
    destinationFileName = d+fileExtension #exemplo '2019February09.xlsx'
    print('>A criar ficheiro-resumo ', destinationFileName, '...')
    # cria o ficheiro
    wr = Workbook()
    wr.save(destinationFileName)

    for file in filesList:
        name = file+'.xlsx'
        wb1 = xl.load_workbook(filename=name) # ficheiro de origem
        ws1 = wb1.worksheets[0]
        wb2 = xl.load_workbook(filename=destinationFileName) #ficheiro de destino
        ws2 = wb2.create_sheet(file)
        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value
        wb2.save(destinationFileName)
    print('>Criado ficheiro ', name, ' com sucesso.')

def fileSender():
    print('>A copiar os ficheiros para a pasta de output...')
    destinationFileName = d + fileExtension  # exemplo '2019February09.xlsx'
    finalPath = defaultPath+d+'\\'
    finalfile = finalPath + destinationFileName
    os.mkdir(finalPath)
    os.rename(destinationFileName, finalfile)

    for file in filesList:
        thisFileDestination = defaultPath + d +'\\' + file + fileExtension
        thisFileOrigin = file + fileExtension
        os.rename(thisFileOrigin, thisFileDestination)
    print('>\n>Cópia dos ficheiros para a pasta de output concluída. ')
    return

def printi(): # função de teste
    print("printi")

def checkPetPrices():
    Petcity.getPetcityPrice()
    Petcity.xlsxPetcity()
    #Amascotados.run()
    return


def checkRunnablePetPrices():
    t1 = threading.Thread(target=Meganimal.run)
    t2 = threading.Thread(target=Miscota.run)
    t3 = threading.Thread(target=Newpetclub.run)
    t4 = threading.Thread(target=Petness.run)
    t5 = threading.Thread(target=Petsonic.run)
    t6 = threading.Thread(target=PiensosYMascotas.run)
    t7 = threading.Thread(target=PlanetaHuerto.run)
    t8 = threading.Thread(target=TiendaAnimal.run)
    t9 = threading.Thread(target=Wecanimal.run)
    t10 = threading.Thread(target=Petcity.run)
    t1.start()
    threads.append(t1)
    time.sleep(15)
    t2.start()
    threads.append(t2)
    time.sleep(15)
    t3.start()
    threads.append(t3)
    time.sleep(15)
    t4.start()
    threads.append(t4)
    time.sleep(15)
    t5.start()
    threads.append(t5)
    time.sleep(15)
    t6.start()
    threads.append(t6)
    time.sleep(15)
    t7.start()
    threads.append(t7)
    time.sleep(15)
    t8.start()
    threads.append(t8)
    time.sleep(15)
    t9.start()
    threads.append(t9)
    t10.start()
    threads.append(t10)

    for thread in threads:
        thread.join()
    return

def checkNOTRunnablePetPrices():
    t1 = threading.Thread(target=Amascotados.run)
    t2 = threading.Thread(target=Kiwoko.run)
    t3 = threading.Thread(target=Biscoitinho.run)
    t4 = threading.Thread(target=Goldpet.run)
    t5 = threading.Thread(target=Zooshop.run)

    t1.start()
    ffthreads.append(t1)
    t2.start()
    ffthreads.append(t2)
    t1.join()
    t2.join()
    t3.start()
    ffthreads.append(t3)
    t4.start()
    ffthreads.append(t4)
    time.sleep(30)
    t5.start()
    ffthreads.append(t5)

    for thread in ffthreads:
        thread.join()


"""
print("d2 =", d)
print(type(d))
total = defaultPath+d
print(total)
finalfile = total+'\\teste.xlsx'
os.rename('teste.xlsx', finalfile)
"""
#print(finalfile)
#cria nova pasta
#os.mkdir(total)


if __name__ == '__main__':
    start_time = time.time()
    print('>Programa a iniciar...')
    checkNOTRunnablePetPrices()
    notRunnablElapsed = time.time() - start_time
    #checkPetPrices()
    start_time2 = time.time()
    checkRunnablePetPrices()
    runnablElapsed = time.time() - start_time2
    elapsed = time.time() - start_time
    print('\n>Duração das que exigem firefox: ', strftime("%M:%S", gmtime(notRunnablElapsed)), 'minutos.')
    print('\n>Duração das tarefas paralelizadas sem firefox: ', strftime("%M:%S", gmtime(runnablElapsed)), 'minutos.')
    print('\n>Programa concluído. Duração total: ', strftime("%M:%S", gmtime(elapsed)), 'minutos.')

    xlsxJoiner()
    fileSender()

    print('\n>>>>>>>FIM<<<<<<<')

    #xlsxJoiner()
    #fileSender()
